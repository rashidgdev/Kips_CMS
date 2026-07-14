"""
Bulk-creates Student/Teacher accounts from an uploaded .xlsx file, instead of
adding people one at a time through the "Add Student"/"Add Teacher" forms.
Each row is validated and created independently (its own transaction), so
one bad row doesn't block the rest of the batch - the caller gets back a
list of what succeeded (with temp passwords) and what failed (with why).
"""
import datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.utils.crypto import get_random_string

from .models import Department, Roles, StaffProfile, StudentProfile, TeacherProfile, User

STUDENT_COLUMNS = [
    ('username', True, 'Unique login name, e.g. ali.hassan'),
    ('first_name', True, ''),
    ('last_name', True, ''),
    ('email', True, 'Must be unique'),
    ('phone_number', False, ''),
    ('roll_number', True, 'Must be unique, e.g. BSCS-2026-014'),
    ('program_code', True, 'Must match an existing Program code, e.g. BSCS'),
    ('semester_number', False, 'e.g. 1 - matched against that program\'s semesters'),
    ('admission_date', True, 'YYYY-MM-DD'),
    ('date_of_birth', False, 'YYYY-MM-DD'),
    ('gender', False, 'male / female / other'),
    ('address', False, ''),
    ('guardian_name', False, ''),
    ('guardian_phone', False, ''),
    ('cnic', False, ''),
]

TEACHER_COLUMNS = [
    ('username', True, 'Unique login name'),
    ('first_name', True, ''),
    ('last_name', True, ''),
    ('email', True, 'Must be unique'),
    ('phone_number', False, ''),
    ('role', False, 'teacher or hod (default: teacher)'),
    ('employee_id', True, 'Must be unique'),
    ('department_code', True, 'Must match an existing Department code'),
    ('designation', False, ''),
    ('qualification', False, ''),
    ('joining_date', True, 'YYYY-MM-DD'),
    ('per_lecture_rate', False, 'Number, default 0'),
]


class RowError(Exception):
    def __init__(self, message):
        self.message = message
        super().__init__(message)


def _norm_header(value):
    return str(value or '').strip().lower().replace(' ', '_')


def read_rows(uploaded_file, columns):
    """Yields (row_number, {field: raw_value}) for each non-blank data row.
    Raises RowError immediately if required headers are missing."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=uploaded_file, read_only=True, data_only=True)
    sheet = wb.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise RowError('The file is empty.')

    headers = [_norm_header(h) for h in header_row]
    required = [name for name, is_required, _ in columns if is_required]
    missing = [name for name in required if name not in headers]
    if missing:
        raise RowError(f'Missing required column(s): {", ".join(missing)}')

    for row_num, raw_row in enumerate(rows_iter, start=2):
        if raw_row is None or all(cell in (None, '') for cell in raw_row):
            continue
        row = dict(zip(headers, raw_row))
        yield row_num, row


def _clean_str(value):
    if value is None:
        return ''
    return str(value).strip()


def _require(row, field, row_num):
    value = _clean_str(row.get(field))
    if not value:
        raise RowError(f'Row {row_num}: "{field}" is required.')
    return value


def _require_raw(row, field, row_num):
    """Like `_require`, but returns the original cell value (e.g. a `datetime`
    object for a date-typed Excel cell) instead of coercing it to a string."""
    value = row.get(field)
    if value is None or (isinstance(value, str) and not value.strip()):
        raise RowError(f'Row {row_num}: "{field}" is required.')
    return value


def _get_by_code(model, code, row_num, label):
    """Looks up `model` by its unique `code` field, tried case-sensitively
    first (codes like "ENG" and "Eng" can legitimately both exist), falling
    back to a case-insensitive match only when that resolves to exactly one
    row - otherwise raises a clear ambiguity error instead of an unhandled
    MultipleObjectsReturned."""
    try:
        return model.objects.get(code=code)
    except model.DoesNotExist:
        pass
    matches = list(model.objects.filter(code__iexact=code))
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        raise RowError(
            f'Row {row_num}: "{code}" matches multiple {label} codes '
            f'({", ".join(m.code for m in matches)}) - use the exact code (case-sensitive).'
        )
    raise RowError(f'Row {row_num}: no {label} found with code "{code}".')


def _parse_date(value, field, row_num):
    if value is None or value == '':
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise RowError(f'Row {row_num}: "{field}" value "{text}" is not a valid date (use YYYY-MM-DD).')


def _parse_decimal(value, field, row_num, default=Decimal('0')):
    if value in (None, ''):
        return default
    try:
        return Decimal(str(value))
    except InvalidOperation:
        raise RowError(f'Row {row_num}: "{field}" value "{value}" is not a valid number.')


def _create_user(username, first_name, last_name, email, phone_number, role):
    temp_password = get_random_string(10)
    user = User.objects.create_user(
        username=username,
        first_name=first_name,
        last_name=last_name,
        email=email,
        phone_number=phone_number,
        role=role,
        must_change_password=True,
        password=temp_password,
    )
    return user, temp_password


def import_students(uploaded_file):
    """Returns (created: [{row, username, roll_number, name, password}], errors: [{row, message}])."""
    created, errors = [], []
    seen_usernames, seen_emails, seen_rolls = set(), set(), set()

    try:
        rows = list(read_rows(uploaded_file, STUDENT_COLUMNS))
    except RowError as exc:
        return [], [{'row': '-', 'message': exc.message}]

    from apps.academics.models import Program, Semester

    for row_num, row in rows:
        try:
            username = _require(row, 'username', row_num)
            first_name = _require(row, 'first_name', row_num)
            last_name = _require(row, 'last_name', row_num)
            email = _require(row, 'email', row_num)
            roll_number = _require(row, 'roll_number', row_num)
            program_code = _require(row, 'program_code', row_num)
            admission_date = _parse_date(_require_raw(row, 'admission_date', row_num), 'admission_date', row_num)

            if username in seen_usernames or User.objects.filter(username=username).exists():
                raise RowError(f'Row {row_num}: username "{username}" is already taken.')
            if email.lower() in seen_emails or User.objects.filter(email__iexact=email).exists():
                raise RowError(f'Row {row_num}: email "{email}" is already in use.')
            if roll_number in seen_rolls or StudentProfile.objects.filter(roll_number=roll_number).exists():
                raise RowError(f'Row {row_num}: roll number "{roll_number}" is already in use.')

            program = _get_by_code(Program, program_code, row_num, 'program')

            semester = None
            semester_number = _clean_str(row.get('semester_number'))
            if semester_number:
                semester = Semester.objects.filter(program=program, number=semester_number).first()
                if semester is None:
                    raise RowError(
                        f'Row {row_num}: no semester {semester_number} found for program "{program_code}".'
                    )

            gender = _clean_str(row.get('gender')).lower()
            if gender and gender not in dict(StudentProfile.Gender.choices):
                raise RowError(f'Row {row_num}: gender "{gender}" must be one of male/female/other.')

            date_of_birth = _parse_date(row.get('date_of_birth'), 'date_of_birth', row_num)

            with transaction.atomic():
                user, temp_password = _create_user(
                    username, first_name, last_name, email, _clean_str(row.get('phone_number')), Roles.STUDENT
                )
                StudentProfile.objects.create(
                    user=user,
                    roll_number=roll_number,
                    program=program,
                    current_semester=semester,
                    admission_date=admission_date,
                    date_of_birth=date_of_birth,
                    gender=gender,
                    address=_clean_str(row.get('address')),
                    guardian_name=_clean_str(row.get('guardian_name')),
                    guardian_phone=_clean_str(row.get('guardian_phone')),
                    cnic=_clean_str(row.get('cnic')),
                )

            seen_usernames.add(username)
            seen_emails.add(email.lower())
            seen_rolls.add(roll_number)
            created.append(
                {
                    'row': row_num, 'username': username, 'name': f'{first_name} {last_name}',
                    'identifier': roll_number, 'password': temp_password,
                }
            )
        except RowError as exc:
            errors.append({'row': row_num, 'message': exc.message})
        except Exception as exc:  # noqa: BLE001 - surface unexpected row errors without failing the batch
            errors.append({'row': row_num, 'message': f'Unexpected error: {exc}'})

    return created, errors


def import_teachers(uploaded_file):
    """Returns (created: [{row, username, employee_id, name, password}], errors: [{row, message}])."""
    created, errors = [], []
    seen_usernames, seen_emails, seen_ids = set(), set(), set()

    try:
        rows = list(read_rows(uploaded_file, TEACHER_COLUMNS))
    except RowError as exc:
        return [], [{'row': '-', 'message': exc.message}]

    for row_num, row in rows:
        try:
            username = _require(row, 'username', row_num)
            first_name = _require(row, 'first_name', row_num)
            last_name = _require(row, 'last_name', row_num)
            email = _require(row, 'email', row_num)
            employee_id = _require(row, 'employee_id', row_num)
            department_code = _require(row, 'department_code', row_num)
            joining_date = _parse_date(_require_raw(row, 'joining_date', row_num), 'joining_date', row_num)

            role_value = _clean_str(row.get('role')).lower() or Roles.TEACHER
            if role_value not in (Roles.TEACHER, Roles.HOD):
                raise RowError(f'Row {row_num}: role "{role_value}" must be "teacher" or "hod".')

            if username in seen_usernames or User.objects.filter(username=username).exists():
                raise RowError(f'Row {row_num}: username "{username}" is already taken.')
            if email.lower() in seen_emails or User.objects.filter(email__iexact=email).exists():
                raise RowError(f'Row {row_num}: email "{email}" is already in use.')
            if employee_id in seen_ids or TeacherProfile.objects.filter(employee_id=employee_id).exists():
                raise RowError(f'Row {row_num}: employee ID "{employee_id}" is already in use.')

            department = _get_by_code(Department, department_code, row_num, 'department')

            per_lecture_rate = _parse_decimal(row.get('per_lecture_rate'), 'per_lecture_rate', row_num)

            with transaction.atomic():
                user, temp_password = _create_user(
                    username, first_name, last_name, email, _clean_str(row.get('phone_number')), role_value
                )
                TeacherProfile.objects.create(
                    user=user,
                    employee_id=employee_id,
                    department=department,
                    designation=_clean_str(row.get('designation')),
                    qualification=_clean_str(row.get('qualification')),
                    joining_date=joining_date,
                    per_lecture_rate=per_lecture_rate,
                )

            seen_usernames.add(username)
            seen_emails.add(email.lower())
            seen_ids.add(employee_id)
            created.append(
                {
                    'row': row_num, 'username': username, 'name': f'{first_name} {last_name}',
                    'identifier': employee_id, 'password': temp_password,
                }
            )
        except RowError as exc:
            errors.append({'row': row_num, 'message': exc.message})
        except Exception as exc:  # noqa: BLE001
            errors.append({'row': row_num, 'message': f'Unexpected error: {exc}'})

    return created, errors


def build_template_workbook(columns):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    sheet = wb.active
    headers = [name for name, _, _ in columns]
    notes = [note for _, _, note in columns]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.append(notes)
    for cell in sheet[2]:
        cell.font = Font(italic=True, color='9CA3AF')
    for i, _ in enumerate(headers, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = 22
    return wb
