"""
Shared Excel/PDF export helpers used across reporting views (daybook
workload, finance fee summary, and the reports app) so every module exports
in the same format instead of hand-rolling its own.
"""
import io
import os

BRAND_BLUE = '#1E3A8A'


def _logo_path():
    from django.conf import settings

    return os.path.join(settings.BASE_DIR, 'static', 'images', 'kips-logo-icon.png')


def draw_pdf_letterhead(c, page_width, page_top, subtitle=''):
    """Draws the KIPS logo + institution name (and optional subtitle) at the
    top of a reportlab page, in the same visual style already used by
    apps/finance/challan_pdf.py and apps/reports/progress_pdf.py. Meant to be
    called once per page (e.g. from a SimpleDocTemplate's onFirstPage/
    onLaterPages callback) so it repeats across every page of a multi-page
    document. Returns the y-coordinate immediately below the letterhead, for
    callers that draw the rest of the page by hand."""
    from reportlab.lib import colors
    from reportlab.lib.units import mm

    left = 12 * mm
    y = page_top - 10 * mm

    logo_path = _logo_path()
    if os.path.exists(logo_path):
        try:
            c.drawImage(
                logo_path, left, y - 10 * mm, width=12 * mm, height=12 * mm,
                preserveAspectRatio=True, mask='auto',
            )
        except Exception:
            pass

    c.setFont('Helvetica-Bold', 13)
    c.setFillColor(colors.HexColor(BRAND_BLUE))
    c.drawString(left + 16 * mm, y - 3 * mm, 'KIPS College Kasur Campus')
    if subtitle:
        c.setFont('Helvetica', 8.5)
        c.setFillColor(colors.HexColor('#4B5563'))
        c.drawString(left + 16 * mm, y - 8 * mm, subtitle)

    c.setStrokeColor(colors.HexColor('#D1D5DB'))
    c.line(left, y - 13 * mm, page_width - left, y - 13 * mm)
    return y - 16 * mm


LETTERHEAD_TOP_MARGIN_MM = 34  # clears draw_pdf_letterhead's ~26mm plus breathing room


def export_excel(filename, headers, rows):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active

    last_col = get_column_letter(max(len(headers), 1))
    sheet.merge_cells(f'A1:{last_col}1')
    sheet['A1'] = 'KIPS College Kasur Campus'
    sheet['A1'].font = Font(bold=True, size=14, color=BRAND_BLUE.lstrip('#'))
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[1].height = 32

    logo_path = _logo_path()
    if os.path.exists(logo_path):
        logo = XLImage(logo_path)
        logo.width = logo.height = 34
        sheet.add_image(logo, 'A1')

    header_row = 3
    for col_index, value in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col_index, value=value)
        cell.font = Font(bold=True)
    for row_offset, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=header_row + row_offset, column=col_index, value=str(value))

    for col_index in range(1, len(headers) + 1):
        letter = get_column_letter(col_index)
        length = max(
            (len(str(sheet.cell(row=r, column=col_index).value or ''))
             for r in range(header_row, sheet.max_row + 1)),
            default=10,
        )
        sheet.column_dimensions[letter].width = min(length + 2, 40)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def export_pdf(filename, title, headers, rows, subtitle=''):
    from django.http import HttpResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4), title=title, topMargin=LETTERHEAD_TOP_MARGIN_MM * mm,
    )
    styles = getSampleStyleSheet()

    def _letterhead(c, doc):
        draw_pdf_letterhead(c, doc.pagesize[0], doc.pagesize[1])

    elements = [Paragraph(title, styles['Title'])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles['Normal']))
    elements.append(Spacer(1, 12))

    data = [headers] + [[str(value) for value in row] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
        )
    )
    elements.append(table)
    doc.build(elements, onFirstPage=_letterhead, onLaterPages=_letterhead)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response
